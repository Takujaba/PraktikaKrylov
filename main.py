import datetime
from io import BytesIO
import time
from typing import Any
from flask import Flask, jsonify, render_template, request, send_file
from ultralytics import YOLO
from ultralytics.utils.plotting import Annotator
from PIL import Image
import numpy as np
import base64
import cv2
import os
from db import DB
import openpyxl


app = Flask(__name__, static_folder='./static')
model = YOLO('./yolov8m-sheep.pt')
db = DB('./db.sqlite')

if not os.path.exists('./images'):
    os.makedirs('./images')

if not os.path.exists('./reports'):
    os.makedirs('./reports')


def create_report(logs: list[Any], report_name: str):
    columns = ['datetime', 'photo_name', 'prediction_time', 'sheep_count']
    wb = openpyxl.Workbook()
    sheet = wb.active

    for row in sheet.iter_rows(min_row=1, max_col=len(columns), max_row=1):
        for idx, cell in enumerate(row):
            cell.value = columns[idx]

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, max_col=len(columns), max_row=len(logs) + 1)
    ):
        for col_idx, cell in enumerate(row):
            cell.value = logs[row_idx][col_idx]

    wb.save(f'./reports/{report_name}.xlsx')
    wb.close()


@app.get('/')
def get_page():
    return render_template('index.html')


@app.post('/')
def post_image():
    img = np.ascontiguousarray(
        cv2.imdecode(
            np.frombuffer(request.files['image'].read(), np.uint8), cv2.IMREAD_COLOR
        )[:, :, :3]
    )

    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    start = time.time()
    results = model.predict(img)
    prediction_time = time.time() - start

    annotator = Annotator(img)
    sheep_count = 0
    for result in results:
        for box in result.boxes:
            if model.names[int(box.cls)] == 'sheep':
                annotator.box_label(box.xyxy[0])
                sheep_count += 1

    photo_name = f'{time.time()}.jpg'
    Image.fromarray(img).save(f'./images/{photo_name}')

    with db.get_cursor() as cursor:
        insert_query = '''
            INSERT INTO logs
            (datetime, photo_name, prediction_time, sheep_count)
            VALUES (?, ?, ?, ?)
        '''
        cursor.execute(
            insert_query,
            [datetime.datetime.now(), photo_name, prediction_time, sheep_count],
        )

    buffer = BytesIO()
    Image.fromarray(annotator.result()).save(buffer, format='JPEG')
    byte_data = buffer.getvalue()
    base64_encoded = base64.b64encode(byte_data).decode('utf-8')
    uri = 'data:image/jpeg;base64,' + base64_encoded
    return jsonify({'image': uri, 'sheep_count': sheep_count})


@app.get('/report')
def get_report():
    with db.get_cursor() as cursor:
        select_query = '''
            SELECT datetime, photo_name,
            prediction_time, sheep_count
            FROM logs
        '''
        logs = cursor.execute(select_query).fetchall()
    report_name = time.time()
    create_report(logs, report_name)
    return send_file(f'./reports/{report_name}.xlsx', as_attachment=True)


if __name__ == '__main__':
    with db.get_cursor() as cursor:
        create_table_query = '''
            CREATE TABLE IF NOT EXISTS logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                datetime TEXT,
                photo_name TEXT,
                prediction_time INTEGER,
                sheep_count INTEGER
            )
        '''
        cursor.execute(create_table_query)
    app.run(debug=True)
